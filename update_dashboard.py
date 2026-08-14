#!/usr/bin/env python3
"""
update_dashboard.py
Dashboard Financiero Keyword S.A.
-----------------------------------
Lee los archivos fuente desde SharePoint (modo automático) o desde la
carpeta local 'datos/' (modo local como fallback) y actualiza index.html.

Modos de ejecución:
    python3 update_dashboard.py              → SharePoint (automático)
    python3 update_dashboard.py --local      → Carpeta local datos/
    python3 update_dashboard.py --debug-sp   → Diagnóstico SharePoint

Requisitos:
    pip3 install openpyxl msal requests python-dotenv
"""

import io
import json
import re
import sys
import datetime
import argparse
from pathlib import Path
from collections import defaultdict, Counter

# ── ARGUMENTOS
parser = argparse.ArgumentParser(description='Dashboard Financiero Keyword')
parser.add_argument('--local',    action='store_true', help='Usar archivos locales en datos/')
parser.add_argument('--debug-sp', action='store_true', help='Diagnóstico de SharePoint')
ARGS = parser.parse_args()

# ── RUTAS
BASE_DIR  = Path(__file__).parent
DATOS_DIR = BASE_DIR / 'datos'
INDEX_PATH = BASE_DIR / 'index.html'

# ── NOMBRES DE ARCHIVOS (deben coincidir con SharePoint y carpeta local)
ARCHIVOS_REQUERIDOS = [
    'RESULTADOS.xlsx',
    'FACTURACION.xlsx',
    'FACTURAS.xlsx',
    'ASIGNACIONES.xlsx',
    'CXC.xlsx',
    'BANCOS.xls',      # puede estar como .xls o .xlsx
]

# ══════════════════════════════════════════════════════════════════════════════
# CARGA DE ARCHIVOS — SharePoint o Local
# ══════════════════════════════════════════════════════════════════════════════

def cargar_sharepoint() -> dict:
    """
    Conecta a SharePoint via Microsoft Graph API y descarga todos los Excel
    de la carpeta FINANCIERO. Retorna dict {nombre: bytes}.
    """
    try:
        import os
        import requests
        from msal import ConfidentialClientApplication
        from dotenv import load_dotenv
    except ImportError as e:
        raise ImportError(
            f"Dependencia faltante: {e}\n"
            "Ejecutá: pip3 install msal requests python-dotenv"
        )

    # Cargar .env desde la misma carpeta del script
    env_path = BASE_DIR / '.env'
    if not env_path.exists():
        raise FileNotFoundError(
            f"No se encontró el archivo .env en {BASE_DIR}\n"
            "Copiá el .env de Telmo a la misma carpeta que update_dashboard.py"
        )
    from dotenv import load_dotenv
    load_dotenv(env_path)

    TENANT_ID      = os.environ['TENANT_ID']
    CLIENT_ID      = os.environ['CLIENT_ID']
    CLIENT_SECRET  = os.environ['CLIENT_SECRET']
    SHAREPOINT_URL = os.environ['SHAREPOINT_URL']
    CARPETA        = os.environ.get('CARPETA', 'Documentos/FINANCIERO')
    GRAPH_BASE     = 'https://graph.microsoft.com/v1.0'

    print("  🔐 Autenticando con Microsoft...")
    app = ConfidentialClientApplication(
        client_id=CLIENT_ID,
        client_credential=CLIENT_SECRET,
        authority=f"https://login.microsoftonline.com/{TENANT_ID}",
    )
    resultado = app.acquire_token_for_client(
        scopes=["https://graph.microsoft.com/.default"]
    )
    if "access_token" not in resultado:
        raise RuntimeError(
            f"Error de autenticación: {resultado.get('error_description', resultado)}"
        )
    token = resultado["access_token"]
    hdrs  = {"Authorization": f"Bearer {token}", "Accept": "application/json"}

    # Site ID
    hostname  = SHAREPOINT_URL.split("/")[2]
    site_path = "/".join(SHAREPOINT_URL.split("/")[3:])
    r = requests.get(f"{GRAPH_BASE}/sites/{hostname}:/{site_path}", headers=hdrs)
    r.raise_for_status()
    site_id = r.json()["id"]

    # Drive ID (busca "Documentos")
    nombre_biblioteca = CARPETA.split("/")[0]
    r = requests.get(f"{GRAPH_BASE}/sites/{site_id}/drives", headers=hdrs)
    r.raise_for_status()
    drives = r.json().get("value", [])
    if not drives:
        raise FileNotFoundError(
            f"No se encontró ninguna biblioteca de documentos en el sitio SharePoint "
            f"({SHAREPOINT_URL}). Verificá con --debug-sp que el sitio es correcto."
        )
    drive_id = next(
        (d["id"] for d in drives if d["name"].lower() == nombre_biblioteca.lower()),
        drives[0]["id"]
    )

    # Listar archivos en FINANCIERO
    subcarpeta = "/".join(CARPETA.split("/")[1:])
    url_lista = f"{GRAPH_BASE}/sites/{site_id}/drives/{drive_id}/root:/{subcarpeta}:/children"
    r = requests.get(url_lista, headers=hdrs)
    r.raise_for_status()
    items = r.json().get("value", [])

    nombres_esperados = {Path(n).stem.upper() for n in ARCHIVOS_REQUERIDOS}
    archivos_excel = [
        a for a in items
        if a["name"].lower().endswith((".xlsx", ".xls"))
        and Path(a["name"]).stem.upper() in nombres_esperados
    ]

    if not archivos_excel:
        raise FileNotFoundError(
            f"No se encontraron archivos Excel en SharePoint/{CARPETA}\n"
            "Verificá con --debug-sp que la carpeta existe y tiene archivos."
        )

    print(f"  📂 {len(archivos_excel)} archivos encontrados en SharePoint/{CARPETA}")

    # Descargar cada archivo en memoria
    contenidos = {}
    for a in archivos_excel:
        nombre = a["name"]
        url_dl = f"{GRAPH_BASE}/sites/{site_id}/drives/{drive_id}/items/{a['id']}/content"
        r = requests.get(url_dl, headers=hdrs, allow_redirects=True)
        r.raise_for_status()
        contenidos[nombre] = r.content
        kb = len(r.content) // 1024
        print(f"    ✓ {nombre} ({kb} KB)")

    return contenidos


def cargar_local() -> dict:
    """
    Lee los archivos desde la carpeta datos/. Retorna dict {nombre: bytes}.
    """
    if not DATOS_DIR.exists():
        raise FileNotFoundError(f"No existe la carpeta: {DATOS_DIR}")

    contenidos = {}
    for nombre in ARCHIVOS_REQUERIDOS:
        path = DATOS_DIR / nombre
        # BANCOS y FACTURAS pueden estar como .xls o .xlsx
        if not path.exists() and nombre == 'BANCOS.xls':
            path_xlsx = DATOS_DIR / 'BANCOS.xlsx'
            if path_xlsx.exists():
                contenidos['BANCOS.xlsx'] = path_xlsx.read_bytes()
                print(f"    ✓ BANCOS.xlsx (local)")
                continue
        if not path.exists() and nombre == 'FACTURAS.xlsx':
            path_xls = DATOS_DIR / 'FACTURAS.xls'
            if path_xls.exists():
                contenidos['FACTURAS.xls'] = path_xls.read_bytes()
                print(f"    ✓ FACTURAS.xls (local)")
                continue
        # RESULTADOS, FACTURACION y CXC también pueden llegar como .xls viejo
        if not path.exists() and nombre in ('RESULTADOS.xlsx', 'FACTURACION.xlsx', 'CXC.xlsx'):
            path_xls = DATOS_DIR / (nombre[:-1])  # .xlsx -> .xls
            if path_xls.exists():
                contenidos[path_xls.name] = path_xls.read_bytes()
                print(f"    ✓ {path_xls.name} (local)")
                continue
        if path.exists():
            contenidos[nombre] = path.read_bytes()
            print(f"    ✓ {nombre} (local)")
        else:
            print(f"    ⚠️  No encontrado: {nombre}")

    return contenidos


def debug_sharepoint():
    """Diagnóstico completo de la conexión y estructura SharePoint."""
    try:
        import os, requests
        from msal import ConfidentialClientApplication
        from dotenv import load_dotenv
        load_dotenv(BASE_DIR / '.env')
    except ImportError as e:
        print(f"❌ {e} — pip3 install msal requests python-dotenv")
        return

    TENANT_ID      = os.environ['TENANT_ID']
    CLIENT_ID      = os.environ['CLIENT_ID']
    CLIENT_SECRET  = os.environ['CLIENT_SECRET']
    SHAREPOINT_URL = os.environ['SHAREPOINT_URL']
    CARPETA        = os.environ.get('CARPETA', 'Documentos/FINANCIERO')
    GRAPH_BASE     = 'https://graph.microsoft.com/v1.0'

    print(f"\n{'='*55}")
    print("DIAGNÓSTICO SHAREPOINT")
    print(f"{'='*55}")
    print(f"  Sitio:   {SHAREPOINT_URL}")
    print(f"  Carpeta: {CARPETA}")

    app = ConfidentialClientApplication(
        CLIENT_ID, CLIENT_SECRET,
        authority=f"https://login.microsoftonline.com/{TENANT_ID}"
    )
    r = app.acquire_token_for_client(["https://graph.microsoft.com/.default"])
    if "access_token" not in r:
        print(f"\n❌ Error de autenticación: {r.get('error_description')}")
        return
    print("\n  ✅ Token obtenido correctamente")
    tk = r["access_token"]
    hdrs = {"Authorization": f"Bearer {tk}"}

    def get(url):
        resp = requests.get(url, headers=hdrs)
        print(f"  HTTP {resp.status_code}  {url.replace(GRAPH_BASE,'')}")
        if not resp.ok:
            print(f"  ❌ {resp.text[:200]}")
            return {}
        return resp.json()

    hostname  = SHAREPOINT_URL.split("/")[2]
    site_path = "/".join(SHAREPOINT_URL.split("/")[3:])
    site      = get(f"{GRAPH_BASE}/sites/{hostname}:/{site_path}")
    if not site:
        return
    site_id = site["id"]
    print(f"\n  ✅ Site ID: {site_id[:40]}...")

    print("\n  Drives disponibles:")
    drives = get(f"{GRAPH_BASE}/sites/{site_id}/drives").get("value", [])
    for d in drives:
        print(f"    • {d['name']}  ({d['id'][:30]}...)")

    drive_id = drives[0]["id"]
    print("\n  Raíz del drive:")
    raiz = get(f"{GRAPH_BASE}/sites/{site_id}/drives/{drive_id}/root/children").get("value", [])
    for i in raiz:
        tipo = "DIR " if "folder" in i else "FILE"
        print(f"    {tipo}  {i['name']}")

    subcarpeta = "/".join(CARPETA.split("/")[1:])
    print(f"\n  Contenido de /{subcarpeta}:")
    fin = get(
        f"{GRAPH_BASE}/sites/{site_id}/drives/{drive_id}/root:/{subcarpeta}:/children"
    ).get("value", [])
    if not fin:
        print("  ⚠️  Carpeta vacía o no encontrada")
    for i in fin:
        tipo = "DIR " if "folder" in i else "FILE"
        print(f"    {tipo}  {i['name']}  ({i.get('size',0):,} bytes)")
    print()


# ══════════════════════════════════════════════════════════════════════════════
# UTILIDADES
# ══════════════════════════════════════════════════════════════════════════════

def _xls_a_xlsx_bytes(xls_bytes: bytes) -> bytes:
    """Convierte un .xls (formato viejo) a .xlsx en memoria, preservando
    nombres de hojas, valores, fechas y — crítico — formatos de texto/padding
    de ceros (ej: código de cuenta '001' guardado como número 1 con formato
    '000' se preserva como texto '001'). Se usa como fallback transparente
    cuando contabilidad sube archivos en formato antiguo."""
    import xlrd
    from openpyxl import Workbook
    try:
        wb_in = xlrd.open_workbook(file_contents=xls_bytes, formatting_info=True)
        tiene_fmt = True
    except Exception:
        # formatting_info puede fallar con algunos xls; caemos a modo básico
        wb_in = xlrd.open_workbook(file_contents=xls_bytes)
        tiene_fmt = False

    def _fmt_str(cell):
        if not tiene_fmt:
            return ''
        try:
            xf = wb_in.xf_list[cell.xf_index]
            return wb_in.format_map[xf.format_key].format_str or ''
        except Exception:
            return ''

    wb_out = Workbook()
    wb_out.remove(wb_out.active)
    for sheet_name in wb_in.sheet_names():
        ws_in  = wb_in.sheet_by_name(sheet_name)
        ws_out = wb_out.create_sheet(title=sheet_name[:31])
        for i in range(ws_in.nrows):
            fila = []
            for j in range(ws_in.ncols):
                cell = ws_in.cell(i, j)
                v = cell.value
                if cell.ctype == xlrd.XL_CELL_DATE:
                    try:
                        v = xlrd.xldate_as_datetime(v, wb_in.datemode)
                    except Exception:
                        pass
                elif cell.ctype == xlrd.XL_CELL_EMPTY:
                    v = None
                elif cell.ctype == xlrd.XL_CELL_BOOLEAN:
                    v = bool(v)
                elif cell.ctype == xlrd.XL_CELL_NUMBER:
                    fmt = _fmt_str(cell)
                    # Formato texto "@" → convertir número a string entero
                    if fmt == '@' and float(v).is_integer():
                        v = str(int(v))
                    # Formato con relleno de ceros "000", "0000", etc.
                    # → preservar padding como string
                    elif fmt and re.fullmatch(r'0+', fmt) and float(v).is_integer():
                        v = str(int(v)).zfill(len(fmt))
                fila.append(v)
            ws_out.append(fila)
    buf = io.BytesIO()
    wb_out.save(buf)
    return buf.getvalue()


def abrir_workbook(contenidos: dict, nombre: str):
    """Abre un workbook openpyxl desde bytes en memoria.
    Si se pide un .xlsx pero solo existe el .xls, lo convierte
    transparentemente (contabilidad a veces sube el formato viejo)."""
    from openpyxl import load_workbook
    # Acepta BANCOS.xls o BANCOS.xlsx
    if nombre not in contenidos:
        if nombre.endswith('.xlsx'):
            alt = nombre[:-5] + '.xls'
        elif nombre.endswith('.xls'):
            alt = nombre[:-4] + '.xlsx'
        else:
            alt = nombre
        if alt in contenidos:
            # Si pedían .xlsx y solo hay .xls → convertir a xlsx en memoria
            if nombre.endswith('.xlsx') and alt.endswith('.xls'):
                print(f"    ℹ️  {alt} está en formato viejo — convirtiendo a xlsx en memoria")
                contenidos[nombre] = _xls_a_xlsx_bytes(contenidos[alt])
            else:
                nombre = alt
        else:
            raise FileNotFoundError(f"Archivo no disponible: {nombre}")
    return load_workbook(io.BytesIO(contenidos[nombre]), read_only=True, data_only=True)

def abrir_xlrd(contenidos: dict, nombre: str):
    """Abre un workbook xlrd (.xls) desde bytes en memoria."""
    import xlrd
    if nombre not in contenidos:
        raise FileNotFoundError(f"Archivo no disponible: {nombre}")
    return xlrd.open_workbook(file_contents=contenidos[nombre])

def safe_float(v):
    """Convierte a float de forma segura. Acepta números y texto con formato
    de miles/decimales como '2,835.75' o '2.835,75'."""
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        s = str(v).strip().replace('$', '').replace(' ', '')
        # Formato europeo (punto=miles, coma=decimal): '2.835,75'
        if re.search(r'\d\.\d{3},\d', s):
            s = s.replace('.', '').replace(',', '.')
        else:
            # Formato anglosajón (coma=miles, punto=decimal): '2,835.75'
            s = s.replace(',', '')
        return float(s)
    except:
        return 0.0

# ── CONSTANTES
CC_GRUPO = {
    'A1':'COMUNICACION','A10':'COMUNICACION',
    'A2':'ASUNTOS PUBLICOS','A11':'ASUNTOS PUBLICOS','A13':'ASUNTOS PUBLICOS',
}
CC_TIPO = {'A1':'fija','A2':'fija','A10':'especial','A11':'especial','A13':'especial'}
VALID_CC = set(CC_GRUPO.keys())
CC_MAP   = {
    'A5': 'A2',
    # 'A4': CC real usado para "Bono Incentivo" en nómina (5.2.01.005), activo
    # desde sep-2025 hasta hoy (2026), nunca dado de alta como CC válido —
    # se descartaba en silencio. Confirmado con el usuario: pertenece a
    # Comunicación (A1).
    'A4': 'A1',
    # '01': typo puntual de "A1" en la nómina de septiembre 2025 (asiento 124,
    # Rol de pagos Período 202509). Solo aparece esa vez en todo 2025-2026.
    # Confirmado con el usuario.
    '01': 'A1',
}
# CC combinados que contabilidad escribe a mano en FACTURACION (en vez de partir
# la fila en dos con su A1/A2 real) → se fuerza split usando el % de ASIGNACIONES
# del cliente si existe, o 50/50 si el cliente no tiene % definido. Confirmado con
# el usuario 2026-08-08 (facturas 5776 AIBE y 5929 MARESA, CC='A1 Y A2').
CC_SPLIT_MANUAL = {'A1 Y A2'}
EXCL_ACCTS = {'5.2.07.090','5.2.07.091'}  # Solo impuestos de cierre de año
EXCL_DESC  = 'CIERRE DE PERIODO'
NOMBRE_DISPLAY = {
    '1.1.1.02.01':'Banco Pichincha',
    '1.1.1.02.02':'Banco Internacional',
    '1.1.1.02.03':'Bco. Internacional Ahorro',
    '1.1.1.02.05':'Inv. Fondo Óptimo ANEFI',
    '1.1.1.02.08':'Inv. ANEFI Corto Plazo',
    '1.1.1.02.09':'Pichincha Cta. Ahorros Ec+',
}
TIPO_CUENTA = {
    '1.1.1.02.01':'bank','1.1.1.02.02':'bank','1.1.1.02.03':'bank',
    '1.1.1.02.05':'inv','1.1.1.02.08':'inv','1.1.1.02.09':'bank',
}
EXCLUIR_CUENTAS = {'1.1.1.01.01','1.1.1.01.02'}


def gasto_grupo(nombre):
    """
    Clasifica una cuenta de gasto (5.2.x) en su subgrupo analítico.
    Orden de evaluación: nom → inst → viaj → gest → serv → otros
    Las categorías son mutuamente excluyentes; se usa la primera que matchea.

    nom  — Nómina y beneficios sociales (relación de DEPENDENCIA laboral)
    inst — Instalaciones y tecnología de base (costos fijos de operación)
    viaj — Viajes, movilización y representación externa
    gest — Gestión comercial y representación institucional
    serv — Servicios profesionales externos y producción (sin dependencia laboral)
    otros— Todo lo que no encaja en las anteriores
    """
    n = str(nombre).upper()

    # ── NOM: Nómina — relación de dependencia laboral ──────────────────────
    # IMPORTANTE: honorarios/consultores NO van aquí aunque digan "remuneración"
    # Solo cuentas con relación laboral formal (IESS, décimos, fondos de reserva)
    if any(x in n for x in [
        'SUELDO', 'SALARIO', 'REMUNERACION', 'REMUNER',
        'APORTE PATRON', 'APORTE IESS', 'IESS',
        'FONDO DE RESERVA', 'DECIMO TERCERO', 'DECIMO CUARTO',
        'VACACIONES', 'DESAHUCIO', 'DESPIDO', 'JUBILACION PATRONAL',
        'BONIFICACION', 'HORAS EXTRA',
        'CAPACITACION', 'UNIFORMES', 'ALIMENTACION EMPLEADOS',
        'ATENCIONES EMPLEADOS', 'BENEFICIO',
    ]):
        return 'nom'

    # ── INST: Instalaciones y tecnología de base ───────────────────────────
    # Costos fijos de la oficina e infraestructura física/digital permanente.
    # HOSTING y plataformas digitales van a SERV (son servicios contratados,
    # no infraestructura propia). Solo van aquí si son activos/contratos fijos.
    if any(x in n for x in [
        'ARRIENDO', 'ALQUILER OFICINA', 'ALQUILER LOCAL',
        'AGUA', 'ENERGIA ELECTRICA', 'LUZ ',
        'TELEFONO FIJO', 'INTERNET OFICINA', 'INTERNET',
        'CONDOMINIO', 'EXPENSAS',
        'AIRE ACONDICIONADO', 'ALARMA',
        'DEPRECIACION', 'AMORTIZACION',
        'MANTENIMIENTO INST', 'MANTENIMIENTO OFICINA', 'MANTENIMIENTO EQUIPO',
        'SUMINISTROS DE OFICINA', 'UTILES DE OFICINA',
        'CUOTAS GREMIALES', 'CUOTA CAMARA',
    ]):
        return 'inst'

    # ── VIAJ: Viajes y movilización ────────────────────────────────────────
    if any(x in n for x in [
        'VIAJE', 'VIATICO', 'PARQUEADERO', 'MOVILIZACION',
        'TAXI', 'UBER', 'TRANSPORTE', 'PASAJE', 'COMBUSTIBLE',
        'PEAJE', 'HOTEL', 'HOSPEDAJE',
    ]):
        return 'viaj'

    # ── GEST: Gestión comercial e institucional ────────────────────────────
    if any(x in n for x in [
        'GESTION', 'REPRESENTACION', 'PUBLICIDAD', 'PROPAGANDA',
        'ATENCIONES SOCIALES', 'ATENCIONES CLIENTES',
        'INTERNACIONALIZ', 'RELACIONES PUBLICAS', 'RELACIONES INSTITUCIONALES',
        'DONACION', 'AUSPICIO', 'PATROCINIO',
        'REGALO', 'OBSEQUIO',
    ]):
        return 'gest'

    # ── SERV: Servicios profesionales y producción externos ────────────────
    # Esta es la categoría más amplia para una empresa de comunicación.
    # Incluye todo servicio contratado a terceros sin relación de dependencia:
    # consultores, productoras, plataformas digitales, outsourcing, etc.
    if any(x in n for x in [
        # Profesionales externos
        'HONORARIO', 'CONSULTOR', 'ASESOR', 'ASESORIA',
        'CONTABILIDAD', 'AUDITORIA', 'ACTUARIO', 'COMISARIO',
        'LEGAL', 'JURIDICO', 'NOTARIA', 'ABOGADO',
        'RECURSOS HUMANOS', 'HEADHUNTER', 'RECLUTAMIENTO',
        'MEDICO', 'EXAMENES', 'SALUD OCUPACIONAL',
        # Producción de contenido (core del negocio de comunicación)
        'PRODUCCION', 'PRODUCCION AUDIOVISUAL', 'PRODUCCION DIGITAL',
        'VIDEO', 'FOTOGRAFIA', 'FOTO ', 'EDICION', 'ANIMACION',
        'DISEÑO', 'DIAGRAMACION', 'ILUSTRACION',
        'IMPRENTA', 'IMPRESION', 'MATERIAL IMPRESO',
        # Plataformas y servicios digitales
        'HOSTING', 'DOMINIO', 'SERVIDOR',
        'GOOGLE', 'META ', 'FACEBOOK ADS', 'LINKEDIN',
        'ADOBE', 'MICROSOFT 365', 'OFFICE 365',
        'SUSCRIPCION', 'LICENCIA SOFTWARE', 'LICENCIA DIGITAL',
        'SOFTWARE', 'APLICACION', 'PLATAFORMA DIGITAL',
        'PAGINA WEB', 'DESARROLLO WEB', 'DESARROLLO',
        # Servicios operativos externos
        'MONITOREO', 'CLIPPING', 'BOLETIN',
        'LIMPIEZA', 'MENSAJERIA', 'COURIER', 'VALIJA',
        'SEGURIDAD', 'VIGILANCIA',
        'SEGURO', 'CALIFICACION', 'CERTIFICACION',
        'OUTSOURC', 'TERCERIZA',
        'ADMINISTRACION TECNOLOG', 'SOPORTE TECNICO', 'SOPORTE',
        'SERVICIO', 'SERVICOS', 'PRESTACION',
        # Telecomunicaciones (celulares son servicio, no infraestructura fija)
        'CELULAR', 'TELEFONO MOVIL', 'DATOS MOVILES',
        'INTERNET ', 'BANDA ANCHA',
        # Otros servicios contratados
        'ALQUILER EQUIPOS', 'ALQUILER VEHICULO',
    ]):
        return 'serv'

    # ── OTROS ──────────────────────────────────────────────────────────────
    return 'otros'


def resolver_cc(cc_raw, contexto='gasto'):
    if not cc_raw or str(cc_raw).strip() in ('','nan','None'):
        return [('A1',1.0)] if contexto=='gasto' else [('A1',0.5),('A2',0.5)]
    cc = str(cc_raw).strip().upper()
    cc = CC_MAP.get(cc, cc)
    if cc == 'A1 Y A2':
        return [('A1',0.5),('A2',0.5)]
    if cc in VALID_CC:
        return [(cc,1.0)]
    return []


# ══════════════════════════════════════════════════════════════════════════════
# 1. PyG
# ══════════════════════════════════════════════════════════════════════════════
def procesar_pyg(contenidos):
    print("  Leyendo RESULTADOS.xlsx ...")
    wb = abrir_workbook(contenidos, 'RESULTADOS.xlsx')
    ws = wb['Sheet1']
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    acum = defaultdict(lambda: defaultdict(float))
    cuenta_actual = None
    nombre_actual = ''
    rows_ok = 0
    cc_no_reconocido = defaultdict(lambda: {'n': 0, 'monto': 0.0})

    for row in all_rows:
        c0 = str(row[0]).strip() if row[0] is not None else ''
        if (c0.startswith('4.') or c0.startswith('5.')) and row[3] is None:
            cuenta_actual = c0
            nombre_actual = str(row[1]) if row[1] else ''
            continue
        if c0 != '001' or not isinstance(row[3], datetime.datetime):
            continue
        if not cuenta_actual or cuenta_actual in EXCL_ACCTS:
            continue
        desc = str(row[4]).upper() if row[4] else ''
        if EXCL_DESC in desc:
            continue
        fecha = row[3]
        mes, anio = fecha.month, fecha.year
        debe  = safe_float(row[9])
        haber = safe_float(row[10])
        if cuenta_actual.startswith('4.'):
            # Ingresos: haber - debe
            valor = haber - debe
            splits = resolver_cc(row[8], 'ingreso')
            if not splits:
                cc_no_reconocido[str(row[8])]['n'] += 1
                cc_no_reconocido[str(row[8])]['monto'] += valor
            for cc, frac in splits:
                acum[(anio,mes,cc)]['ing'] += valor * frac
                rows_ok += 1
        elif cuenta_actual.startswith('5.1.'):
            # Costos: debe - haber
            valor = debe - haber
            splits = resolver_cc(row[8], 'gasto')
            if not splits:
                cc_no_reconocido[str(row[8])]['n'] += 1
                cc_no_reconocido[str(row[8])]['monto'] += valor
            for cc, frac in splits:
                acum[(anio,mes,cc)]['costo'] += valor * frac
                rows_ok += 1
        elif cuenta_actual.startswith('5.2.'):
            # Gastos: debe - haber
            valor = debe - haber
            grp = gasto_grupo(nombre_actual)
            splits = resolver_cc(row[8], 'gasto')
            if not splits:
                cc_no_reconocido[str(row[8])]['n'] += 1
                cc_no_reconocido[str(row[8])]['monto'] += valor
            for cc, frac in splits:
                acum[(anio,mes,cc)][grp] += valor * frac
                rows_ok += 1

    print(f"    {rows_ok} movimientos procesados")

    if cc_no_reconocido:
        print(f"    ⚠️  Centro de Costo no reconocido (excluido del PyG, revisar):")
        for cc_raw, info in sorted(cc_no_reconocido.items(), key=lambda x: -abs(x[1]['monto'])):
            print(f"       CC={cc_raw!r}  {info['n']} fila(s)  ${info['monto']:,.2f}")

    # ── Audit log: mostrar cuentas que cayeron en 'otros' para revisión ──
    if acum:
        otros_cuentas = defaultdict(float)
        # Re-scan para identificar cuentas en 'otros'
        cuenta_actual_audit = None
        nombre_actual_audit = ''
        for row in all_rows:
            c0 = str(row[0]).strip() if row[0] is not None else ''
            if (c0.startswith('4.') or c0.startswith('5.')) and row[3] is None:
                cuenta_actual_audit = c0
                nombre_actual_audit = str(row[1]) if row[1] else ''
                continue
            if c0 != '001' or not isinstance(row[3], datetime.datetime): continue
            if not cuenta_actual_audit or cuenta_actual_audit in EXCL_ACCTS: continue
            if not cuenta_actual_audit.startswith('5.2.'): continue
            grp = gasto_grupo(nombre_actual_audit)
            if grp == 'otros':
                egr = safe_float(row[10] if len(row) > 10 else 0)
                if egr > 0:
                    otros_cuentas[f"{cuenta_actual_audit} — {nombre_actual_audit}"] += egr
        if otros_cuentas:
            print(f"    ℹ️  Cuentas clasificadas en 'otros' (revisar si corresponde):")
            for nombre_c, monto in sorted(otros_cuentas.items(), key=lambda x: -x[1])[:10]:
                print(f"       ${monto:>10,.2f}  {nombre_c[:70]}")
            if len(otros_cuentas) > 10:
                print(f"       ... y {len(otros_cuentas)-10} cuentas más")

    resultado = defaultdict(list)
    for (anio,mes,cc), vals in sorted(acum.items()):
        nom=round(vals.get('nom',0),2); inst=round(vals.get('inst',0),2)
        viaj=round(vals.get('viaj',0),2); gest=round(vals.get('gest',0),2)
        serv=round(vals.get('serv',0),2); otros=round(vals.get('otros',0),2)
        ing=round(vals.get('ing',0),2);   costo=round(vals.get('costo',0),2)
        gastos=round(nom+inst+viaj+gest+serv+otros,2)
        resultado[anio].append({
            'month':mes,'cc':cc,'grupo':CC_GRUPO.get(cc,cc),
            'tipo_neg':CC_TIPO.get(cc,'fija'),
            'ing':ing,'costo':costo,'mb':round(ing-costo,2),
            'nom':nom,'inst':inst,'viaj':viaj,'gest':gest,'serv':serv,'otros':otros,
            'gastos':gastos,'res':round(ing-costo-gastos,2),
        })
    return dict(resultado)


# ══════════════════════════════════════════════════════════════════════════════
# 2. Facturación
# ══════════════════════════════════════════════════════════════════════════════
def leer_asignaciones(contenidos):
    wb = abrir_workbook(contenidos, 'ASIGNACIONES.xlsx')
    ws = wb.active
    splits = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        cliente, cat, pct = row[0], row[1], row[2]
        if cliente and cat and pct is not None:
            splits.setdefault(cliente, {})[str(cat).strip()] = float(pct)
    wb.close()
    return splits


def _norm_nro(v):
    """Normaliza un nro de asiento/factura a int, sin importar si viene como
    int, float ('5453.0'), str ('5453'), o con espacios. Retorna None si no
    se puede convertir. Crítico para cruzar FACTURAS.xls con FACTURACION.xls
    porque cada uno puede guardar el nro con tipo distinto."""
    if v is None or v == '':
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, int):
        return v
    if isinstance(v, float):
        return int(v) if v.is_integer() else None
    try:
        s = str(v).strip()
        if not s:
            return None
        return int(float(s))
    except (ValueError, TypeError):
        return None


def procesar_facturacion(contenidos):
    print("  Leyendo ASIGNACIONES.xlsx ...")
    asig = leer_asignaciones(contenidos)
    print(f"    {len(asig)} clientes con split")

    # ── Leer FACTURAS: cruce nro→cliente y sets por tipo ──────────────────────
    print("  Leyendo FACTURAS ...")
    cruce    = {}   # nro → nombre cliente
    fact_nros_vta = set()   # nros VTA05+DVT05 emitidos
    fact_nros_vta06 = []    # filas VTA06 completas (reembolsos)

    def _leer_facturas_rows(rows_iter):
        for row in rows_iter:
            tipo = str(row[1]).strip() if row[1] else ''
            nro = _norm_nro(row[2])
            if nro is None: continue
            cliente = str(row[5]).strip() if row[5] else ''
            fecha   = row[3]
            if tipo in ('VTA05', 'DVT05'):
                if cliente: cruce[nro] = cliente
                fact_nros_vta.add(nro)
            elif tipo == 'VTA06':
                total = safe_float(row[9])   # col[9]=total en FACTURAS para VTA06
                if total and isinstance(fecha, datetime.datetime):
                    fact_nros_vta06.append({
                        'nro': nro, 'cliente': cliente,
                        'fecha': fecha, 'total': total,
                    })

    if 'FACTURAS.xlsx' in contenidos:
        wb = abrir_workbook(contenidos, 'FACTURAS.xlsx')
        ws = wb['Sheet1']
        _leer_facturas_rows(ws.iter_rows(min_row=2, values_only=True))
        wb.close()
    elif 'FACTURAS.xls' in contenidos:
        import xlrd
        wb_xls = abrir_xlrd(contenidos, 'FACTURAS.xls')
        ws_xls = wb_xls.sheet_by_index(0)
        def _xlrd_rows(ws_xls, wb_xls):
            for i in range(1, ws_xls.nrows):
                row = list(ws_xls.row_values(i))
                # Convertir fechas xlrd (float) a datetime
                for j in [3]:
                    cell = ws_xls.cell(i, j)
                    if cell.ctype == xlrd.XL_CELL_DATE:
                        try: row[j] = xlrd.xldate_as_datetime(cell.value, wb_xls.datemode)
                        except: pass
                yield row
        _leer_facturas_rows(_xlrd_rows(ws_xls, wb_xls))

    print(f"    {len(cruce)} facturas en cruce  |  {len(fact_nros_vta06)} VTA06 reembolsos")

    # ── Leer FACTURACION (mayor contable) ─────────────────────────────────────
    print("  Leyendo FACTURACION.xlsx ...")
    wb2 = abrir_workbook(contenidos, 'FACTURACION.xlsx')
    ws2 = wb2['Sheet1']
    all_rows = list(ws2.iter_rows(values_only=True))
    wb2.close()

    acum        = defaultdict(float)   # facturación normal
    acum_reem   = defaultdict(float)   # reembolsos VTA06 (separado)
    sin_emitir  = []   # en FACTURACION pero no en FACTURAS → alerta, NO suma
    rows_ok     = 0
    cc_no_reconocido = defaultdict(lambda: {'n': 0, 'monto': 0.0})

    for row in all_rows:
        if row[0] != '001' or not isinstance(row[3], datetime.datetime):
            continue
        tipo    = str(row[1]).strip() if row[1] else ''
        nro_int = _norm_nro(row[2])
        fecha   = row[3]
        cc_orig = str(row[8]).strip() if row[8] else ''
        # DVT05: monto en col[9]; VTA05: monto en col[10]
        importe = safe_float(row[9]) if tipo == 'DVT05' else safe_float(row[10])
        # CC combinado escrito a mano por contabilidad (ej. 'A1 Y A2') → forzar split,
        # igual que un A1/A2 normal con % de ASIGNACIONES (o 50/50 si el cliente no tiene % definido)
        cc_split_manual = cc_orig in CC_SPLIT_MANUAL
        if not cc_split_manual and cc_orig not in VALID_CC and cc_orig not in CC_MAP:
            cc_no_reconocido[cc_orig]['n'] += 1
            cc_no_reconocido[cc_orig]['monto'] += importe
            continue
        if not cc_split_manual:
            cc_orig = CC_MAP.get(cc_orig, cc_orig)
            if cc_orig not in VALID_CC:
                continue
        if nro_int is None:
            continue
        mes, anio = fecha.month, fecha.year

        # ── Regla principal: excluir facturas sin respaldo en FACTURAS ─────
        if nro_int not in fact_nros_vta:
            sin_emitir.append({
                'nro': nro_int, 'tipo': tipo, 'cc': cc_orig,
                'importe': importe, 'mes': mes, 'anio': anio,
            })
            continue   # NO acumular en el total

        cliente  = cruce.get(nro_int, f'DESCONOCIDO-{nro_int}')
        pct_asig = asig.get(cliente, {})

        if cc_split_manual:
            split = pct_asig if pct_asig else {'A1': 0.5, 'A2': 0.5}
            for cc, pct in split.items():
                imp = (importe * -1 * pct) if tipo == 'DVT05' else pct * importe
                acum[(anio, mes, cliente, cc)] += imp
                rows_ok += 1
        # Split SOLO A1/A2 — A10/A11/A13 son proyectos especiales, nunca se dividen
        elif pct_asig and cc_orig in ('A1', 'A2'):
            for cc, pct in pct_asig.items():
                imp = (importe * -1 * pct) if tipo == 'DVT05' else pct * importe
                acum[(anio, mes, cliente, cc)] += imp
                rows_ok += 1
        else:
            imp = (importe * -1) if tipo == 'DVT05' else importe
            acum[(anio, mes, cliente, cc_orig)] += imp
            rows_ok += 1

    # ── Facturas en FACTURAS pero no en FACTURACION → sí suman ───────────────
    fc_nros = set()
    for row in all_rows:
        if row[0] == '001':
            n = _norm_nro(row[2])
            if n is not None:
                fc_nros.add(n)

    solo_en_facturas = 0

    # Recorrer FACTURAS para filas VTA05/DVT05 que no están en FACTURACION
    def _acum_solo_facturas(rows_iter):
        nonlocal solo_en_facturas
        for row in rows_iter:
            tipo = str(row[1]).strip() if row[1] else ''
            if tipo not in ('VTA05', 'DVT05'): continue
            nro_int = _norm_nro(row[2])
            if nro_int is None: continue
            if nro_int in fc_nros: continue   # ya está en FACTURACION
            fecha = row[3]
            if not isinstance(fecha, datetime.datetime): continue
            cliente = str(row[5]).strip() if row[5] else f'DESCONOCIDO-{nro_int}'
            base    = safe_float(row[7])   # col[7]=base en FACTURAS
            if base == 0: continue
            mes, anio = fecha.month, fecha.year
            # Sin CC propio — usar A2 como default (Asuntos Públicos, caso PLUMATEX)
            cc = 'A2'
            imp = (-base if tipo == 'DVT05' else base)
            acum[(anio, mes, cliente, cc)] += imp
            solo_en_facturas += 1

    if 'FACTURAS.xlsx' in contenidos:
        wb = abrir_workbook(contenidos, 'FACTURAS.xlsx')
        ws = wb['Sheet1']
        _acum_solo_facturas(ws.iter_rows(min_row=2, values_only=True))
        wb.close()
    elif 'FACTURAS.xls' in contenidos:
        import xlrd
        wb_xls = abrir_xlrd(contenidos, 'FACTURAS.xls')
        ws_xls = wb_xls.sheet_by_index(0)
        def _xlrd_rows2(ws_xls, wb_xls):
            for i in range(1, ws_xls.nrows):
                row = list(ws_xls.row_values(i))
                for j in [3]:
                    cell = ws_xls.cell(i, j)
                    if cell.ctype == xlrd.XL_CELL_DATE:
                        try: row[j] = xlrd.xldate_as_datetime(cell.value, wb_xls.datemode)
                        except: pass
                yield row
        _acum_solo_facturas(_xlrd_rows2(ws_xls, wb_xls))

    # ── VTA06 reembolsos ──────────────────────────────────────────────────────
    for r in fact_nros_vta06:
        mes, anio = r['fecha'].month, r['fecha'].year
        cliente   = r['cliente'] or f'DESCONOCIDO-VTA06-{r["nro"]}'
        # Reembolsos van a A10 en acumulador separado
        acum_reem[(anio, mes, cliente, 'A10')] += r['total']

    # ── Resumen consola ───────────────────────────────────────────────────────
    print(f"    {rows_ok} registros de FACTURACION procesados")
    print(f"    {solo_en_facturas} registros solo en FACTURAS añadidos")
    print(f"    {len(fact_nros_vta06)} reembolsos VTA06 añadidos")
    if sin_emitir:
        print(f"    ⚠️  {len(sin_emitir)} fila(s) en FACTURACION sin emisión en FACTURAS (excluidas del total):")
        for s in sin_emitir:
            print(f"       nro={s['nro']}  tipo={s['tipo']}  cc={s['cc']}  "
                  f"importe={s['importe']:,.2f}  {s['anio']}-{s['mes']:02d}")
    if cc_no_reconocido:
        print(f"    ⚠️  Centro de Costo no reconocido (excluido de Facturación, revisar):")
        for cc_raw, info in sorted(cc_no_reconocido.items(), key=lambda x: -abs(x[1]['monto'])):
            print(f"       CC={cc_raw!r}  {info['n']} fila(s)  ${info['monto']:,.2f}")

    resultado = defaultdict(list)
    # Facturación normal
    for (anio, mes, cliente, cc), valor in sorted(acum.items()):
        resultado[anio].append({
            'cliente':   cliente,
            'cc':        cc,
            'grupo':     CC_GRUPO.get(cc, 'OTROS'),
            'tipo':      CC_TIPO.get(cc, 'fija'),
            'mes':       mes,
            'valor':     round(valor, 2),
            'alerta':    cliente.startswith('DESCONOCIDO'),
            'reembolso': False,
        })
    # Reembolsos VTA06
    for (anio, mes, cliente, cc), valor in sorted(acum_reem.items()):
        resultado[anio].append({
            'cliente':   cliente,
            'cc':        cc,
            'grupo':     CC_GRUPO.get(cc, 'OTROS'),
            'tipo':      CC_TIPO.get(cc, 'fija'),
            'mes':       mes,
            'valor':     round(valor, 2),
            'alerta':    cliente.startswith('DESCONOCIDO'),
            'reembolso': True,
        })

    # Pasar lista de sin_emitir al HTML para mostrar en dashboard
    resultado['_sin_emitir'] = sin_emitir
    return dict(resultado)


# ══════════════════════════════════════════════════════════════════════════════
# 3. CxC
# ══════════════════════════════════════════════════════════════════════════════
def procesar_cxc(contenidos):
    """Lee CXC.xlsx/.xls — reporte 'estado de cuenta por cliente' agrupado
    (formato adoptado por contabilidad desde 2026-08-13, reemplaza el listado
    plano anterior). Columnas: LOC | TIPO | NUM | VEN | F.FACT | F.PAGO | DIAS
    | DEBITO | CREDITO | SALDO.

    El archivo intercala 3 tipos de fila sin ninguna marca explícita:
      - fila de cliente:   LOC empieza con 'C' (código cliente), TIPO=nombre, resto vacío
      - fila de factura:   LOC=='001' literal, F.FACT es fecha
      - fila de subtotal / vendedor / total general: no matchea ninguna de las
        dos anteriores (LOC es número, o '01', o vacío) → se ignora sola.
    'valor' usa SALDO (pendiente tras notas de crédito), no DEBITO (importe
    original) — es lo que consumen el total de CxC y el chequeo de vencidas.
    """
    print("  Leyendo CXC.xlsx ...")
    wb = abrir_workbook(contenidos, 'CXC.xlsx')
    ws = wb.active
    registros = []
    today = datetime.date.today()
    cliente_actual = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        row = (list(row) + [None] * 10)[:10]
        loc, tipo, num, ven, ffact, fpago, dias, debito, credito, saldo = row

        if isinstance(loc, str) and loc.strip().startswith('C') and isinstance(tipo, str) and not isinstance(ffact, datetime.datetime):
            cliente_actual = tipo.strip()
            continue

        if loc == '001' and isinstance(ffact, datetime.datetime) and num:
            fem_d  = ffact.date()
            fven_d = fpago.date() if isinstance(fpago, datetime.datetime) else None
            valor = saldo if isinstance(saldo, (int, float)) else safe_float(saldo)
            estado = 'Vencida' if fven_d and fven_d < today else 'Futura'
            registros.append({
                'no': str(num),
                'fem': fem_d.strftime('%d/%m/%Y'),
                'fven': fven_d.strftime('%d/%m/%Y') if fven_d else '',
                'fem_iso': fem_d.isoformat(),
                'fven_iso': fven_d.isoformat() if fven_d else '',
                'cliente': cliente_actual or '',
                'valor': round(float(valor), 2),
                'estado': estado,
                'mes_emision': fem_d.month,
                'mes_vencimiento': fven_d.month if fven_d else 0,
            })
    wb.close()
    if not registros:
        print("    ⚠️  0 registros CxC — el formato del archivo puede haber cambiado de nuevo, revisar estructura de columnas")
    print(f"    {len(registros)} registros CxC")
    return registros


# ══════════════════════════════════════════════════════════════════════════════
# 4. Bancos
# ══════════════════════════════════════════════════════════════════════════════
def procesar_bancos(contenidos):
    # Preferir BANCOS.xlsx si existe, si no usar BANCOS.xls via xlrd
    if 'BANCOS.xlsx' in contenidos:
        print("  Leyendo BANCOS.xlsx ...")
        wb = abrir_workbook(contenidos, 'BANCOS.xlsx')
        ws = wb['Sheet1']
        all_rows = list(ws.iter_rows(values_only=True))
        wb.close()
    elif 'BANCOS.xls' in contenidos:
        print("  Leyendo BANCOS.xls ...")
        import xlrd
        wb_xls = abrir_xlrd(contenidos, 'BANCOS.xls')
        ws_xls = wb_xls.sheet_by_index(0)
        all_rows = []
        for i in range(ws_xls.nrows):
            row = []
            for j in range(ws_xls.ncols):
                cell = ws_xls.cell(i, j)
                if cell.ctype == xlrd.XL_CELL_DATE:
                    try: row.append(xlrd.xldate_as_datetime(cell.value, wb_xls.datemode))
                    except: row.append(cell.value)
                elif cell.ctype == xlrd.XL_CELL_EMPTY:
                    row.append(None)
                else:
                    row.append(cell.value)
            all_rows.append(tuple(row))
        print(f"    {len(all_rows)} filas leídas")
    else:
        print("  ❌ No se encontró BANCOS.xls ni BANCOS.xlsx")
        return [], [], []

    # Detectar subcuentas
    subcuentas = []
    for i, row in enumerate(all_rows):
        c0 = str(row[0]).strip() if row[0] is not None else ''
        val11 = row[11] if len(row) > 11 else None
        if c0.startswith('1.1.1') and isinstance(val11,(int,float)):
            subcuentas.append({'codigo':c0,'nombre':str(row[1]).strip() if row[1] else '','si':float(val11),'fila':i})
    for idx, sc in enumerate(subcuentas):
        sc['fila_fin'] = subcuentas[idx+1]['fila']-1 if idx+1<len(subcuentas) else len(all_rows)-1

    saldo_data = []
    monthly_acum = defaultdict(lambda:{'ing':0.0,'egr':0.0})
    mexico_acum  = defaultdict(float)

    for sc in subcuentas:
        codigo = sc['codigo']
        if codigo in EXCLUIR_CUENTAS:
            continue
        ing_total, egr_total, last_bal = 0.0, 0.0, None
        for row in all_rows[sc['fila']+1 : sc['fila_fin']+1]:
            c0 = str(row[0]).strip() if row[0] is not None else ''
            if c0 != '001': continue
            fecha = row[3] if len(row)>3 else None
            if not isinstance(fecha, datetime.datetime): continue
            cc_raw = str(row[8]).strip() if len(row)>8 and row[8] else ''
            ing  = safe_float(row[9]  if len(row)>9  else 0)
            egr  = safe_float(row[10] if len(row)>10 else 0)
            bal  = row[11] if len(row)>11 else None
            mes  = fecha.month
            if isinstance(bal,(int,float)):
                last_bal = float(bal)
            if cc_raw.upper() == 'MEXICO':
                mexico_acum[mes] += egr
                egr_total += egr
                continue
            monthly_acum[mes]['ing'] += ing
            monthly_acum[mes]['egr'] += egr
            ing_total += ing
            egr_total += egr

        sf = last_bal if last_bal is not None else sc['si']+ing_total-egr_total
        saldo_data.append({
            'banco': NOMBRE_DISPLAY.get(codigo, sc['nombre']),
            'type':  TIPO_CUENTA.get(codigo,'bank'),
            'si':    round(sc['si'],2),
            'ing':   round(ing_total,2),
            'egr':   round(egr_total,2),
            'sf':    round(sf,2),
        })

    monthly = [{'month':m,'ing':round(v['ing'],2),'egr':round(v['egr'],2)}
               for m,v in sorted(monthly_acum.items())]
    mexico  = [{'mes':m,'egr':round(v,2)} for m,v in sorted(mexico_acum.items())]

    print(f"    {len(saldo_data)} subcuentas | {len(monthly)} meses | {len(mexico)} meses México")
    return saldo_data, monthly, mexico


# ══════════════════════════════════════════════════════════════════════════════
# 5. Actualizar HTML
# ══════════════════════════════════════════════════════════════════════════════
def actualizar_html(pyg_data, fac_data, cxc_data, saldo_data, monthly, mexico):
    print("  Actualizando index.html ...")
    if not INDEX_PATH.exists():
        raise FileNotFoundError(f"No se encontró: {INDEX_PATH}")

    with open(INDEX_PATH,'r',encoding='utf-8') as f:
        html = f.read()
    # Backup
    with open(INDEX_PATH.with_suffix('.html.bak'),'w',encoding='utf-8') as f:
        f.write(html)

    meses_2026 = sorted(set(r['month'] for r in pyg_data.get(2026,[])))
    # Separar alertas de sin_emitir antes de serializar
    sin_emitir = fac_data.pop('_sin_emitir', [])
    fac_meses  = sorted(set(r['mes']   for r in fac_data.get(2026,[])))

    # ── Preservar datos 2025 hardcodeados en el HTML existente ──────────────
    # El script solo procesa datos del año en curso (2026). Los datos de 2025
    # están hardcodeados en el index.html y NO deben sobreescribirse con [].
    def _extract_2025_from_html(html_text, var_name):
        """Extrae el array 2025 de PYG_ALL_RAW o FAC_ALL_RAW del HTML actual."""
        try:
            # Encuentra el bloque de la variable
            start = html_text.find(f'const {var_name}')
            if start == -1:
                return '[]'
            # Encuentra '2025:' dentro del bloque
            p25 = html_text.find('  2025:', start)
            if p25 == -1:
                return '[]'
            arr_start = html_text.index('[', p25)
            depth = 0
            for i in range(arr_start, len(html_text)):
                if html_text[i] == '[': depth += 1
                elif html_text[i] == ']':
                    depth -= 1
                    if depth == 0:
                        candidate = html_text[arr_start:i+1]
                        # Validar que no está vacío
                        parsed = json.loads(candidate)
                        if parsed:
                            return candidate
                        return '[]'
        except Exception:
            pass
        return '[]'

    pyg_2025_str = _extract_2025_from_html(html, 'PYG_ALL_RAW')
    fac_2025_str = _extract_2025_from_html(html, 'FAC_ALL_RAW')

    # Meses 2025 para PYG_MONTHS_BY_YEAR
    try:
        pyg_2025_parsed = json.loads(pyg_2025_str)
        meses_2025 = sorted(set(r['month'] for r in pyg_2025_parsed))
    except Exception:
        meses_2025 = []

    if pyg_2025_str != '[]':
        print(f"    ℹ️  Preservando datos PyG 2025 hardcodeados ({len(json.loads(pyg_2025_str))} registros)")
    if fac_2025_str != '[]':
        print(f"    ℹ️  Preservando datos FAC 2025 hardcodeados ({len(json.loads(fac_2025_str))} registros)")

    # ── Guardia: si no se pudo extraer 2025 del HTML existente, abortar en vez
    # de publicar con el año vacío. El backup .bak ya se escribió arriba, y
    # como la escritura real a INDEX_PATH ocurre al final de esta función,
    # abortar acá no corrompe nada — solo detiene el proceso con un error claro.
    if pyg_2025_str == '[]' or fac_2025_str == '[]':
        raise RuntimeError(
            "No se pudieron extraer los datos 2025 del index.html existente "
            "(PYG_ALL_RAW o FAC_ALL_RAW). Abortando para no sobreescribir 2025 con "
            "un array vacío. Revisar index.html.bak y el formato del bloque '2025:'."
        )

    def rep(pat, nuevo, flags=0):
        nonlocal html
        html, count = re.subn(pat, lambda m: nuevo, html, flags=flags)
        if count == 0:
            raise RuntimeError(
                f"No se encontró el patrón esperado en index.html, abortando sin escribir "
                f"(nada se sobreescribió): {pat}"
            )

    rep(r'const PYG_ALL_RAW = \{.*?\};',
        'const PYG_ALL_RAW = {\n'
        f'  2026:{json.dumps(pyg_data.get(2026,[]),ensure_ascii=False)},\n'
        f'  2025:{pyg_2025_str},\n'
        '  2024:[]\n};', re.DOTALL)

    rep(r'const PYG_MONTHS_BY_YEAR = \{.*?\};',
        f'const PYG_MONTHS_BY_YEAR = {{2026:{json.dumps(meses_2026)}, 2025:{json.dumps(meses_2025)}, 2024:[]}};')

    rep(r'const FAC_ALL_RAW = \{.*?\};',
        'const FAC_ALL_RAW = {\n'
        f'  2026:{json.dumps(fac_data.get(2026,[]),ensure_ascii=False)},\n'
        f'  2025:{fac_2025_str},\n'
        '  2024:[]\n};', re.DOTALL)

    rep(r'const FAC_MONTHS = \[.*?\];', f'const FAC_MONTHS = {json.dumps(fac_meses)};')
    rep(r'const BVR_REAL_MONTHS = \[.*?\];', f'const BVR_REAL_MONTHS = {json.dumps(meses_2026)};')

    # NOTA: `sin_emitir` (facturas en FACTURACION sin respaldo en FACTURAS) se
    # imprime en consola desde procesar_facturacion(). No existe un
    # `FAC_SIN_EMITIR` en index.html (nunca se construyó esa parte visual) —
    # antes había un rep() que intentaba escribirlo y fallaba en silencio en
    # cada corrida sin que nadie lo notara. Se removió; si en el futuro se
    # quiere mostrar en el dashboard, hay que agregar el const en index.html
    # primero y luego el rep() acá.

    rep(r'const CXC=\[.*?\];',
        f'const CXC={json.dumps(cxc_data,ensure_ascii=False)};', re.DOTALL)

    saldo_str = 'const SALDO_DATA=[\n' + ',\n'.join(
        f'  {{banco:{json.dumps(r["banco"])},type:{json.dumps(r["type"])},'
        f'si:{r["si"]},ing:{r["ing"]},egr:{r["egr"]},sf:{r["sf"]}}}'
        for r in saldo_data) + '\n];'
    rep(r'const SALDO_DATA=\[.*?\];', saldo_str, re.DOTALL)

    monthly_str = 'const MONTHLY=[\n' + ',\n'.join(
        f'  {{month:{r["month"]},ing:{r["ing"]},egr:{r["egr"]}}}' for r in monthly) + '\n];'
    rep(r'const MONTHLY=\[.*?\];', monthly_str, re.DOTALL)

    mexico_str = 'const MEXICO_DATA=[\n' + ',\n'.join(
        f'  {{mes:{r["mes"]},egr:{r["egr"]}}}' for r in mexico) + '\n];'
    rep(r'const MEXICO_DATA=\[.*?\];', mexico_str, re.DOTALL)

    # ── Actualizar panel estático de Flujo de Caja (SI / SF / Variación / Fecha) ──
    if saldo_data and monthly:
        import calendar
        si_total   = round(sum(r['si'] for r in saldo_data), 0)
        sf_total   = round(sum(r['sf'] for r in saldo_data), 0)
        variacion  = round(sf_total - si_total, 0)
        ultimo_mes = max(r['month'] for r in monthly)
        ultimo_dia = calendar.monthrange(2026, ultimo_mes)[1]
        fecha_str  = "Al " + f"{ultimo_dia:02d}/{ultimo_mes:02d}/2026"

        def fmt_miles(v):
            if v < 0:
                return "($" + f"{abs(v):,.0f}" + ")"
            return "$" + f"{v:,.0f}"

        def rep_tarjeta(label, valor, subfijo):
            nonlocal html
            marker = '<div class="metric-label">' + label + '</div>'
            idx = html.find(marker)
            if idx == -1:
                return
            cierre = '</div>\n    </div>'
            fin = html.find(cierre, idx)
            if fin == -1:
                return
            fin += len(cierre)
            if "ariaci" in label and valor < 0:
                clase = " neg"
            elif "ariaci" in label and valor > 0:
                clase = " pos"
            else:
                clase = ""
            nuevo = (
                '<div class="metric-label">' + label + '</div>\n' +
                '      <div class="metric-value' + clase + '">' + fmt_miles(valor) + '</div>\n' +
                '      <div class="metric-sub">' + subfijo + '</div>\n' +
                '    </div>'
            )
            html = html[:idx] + nuevo + html[fin:]

        rep_tarjeta('Saldo inicial',  si_total,  fecha_str)
        rep_tarjeta('Saldo final',    sf_total,  fecha_str)
        rep_tarjeta('Variaci\u00f3n neta', variacion, 'SF − SI · cuadrado ✓')

        html = re.sub(
            r'(<div class="page-updated">Actualizado [^<]{1,10})' + r'\d{2}/\d{2}/\d{4}(</div>)',
            lambda m: m.group(1) + f"{ultimo_dia:02d}/{ultimo_mes:02d}/2026" + m.group(2),
            html)
    with open(INDEX_PATH,'w',encoding='utf-8') as f:
        f.write(html)
    print("    ✅ index.html actualizado")


# ══════════════════════════════════════════════════════════════════════════════
# 6. Resumen de validación
# ══════════════════════════════════════════════════════════════════════════════
def imprimir_resumen(pyg_data, fac_data, cxc_data, saldo_data, monthly, mexico):
    sep = "="*55
    print(f"\n{sep}\nRESUMEN DE VALIDACIÓN\n{sep}")
    MESES = ['','Ene','Feb','Mar','Abr','May','Jun','Jul','Ago','Sep','Oct','Nov','Dic']
    alertas = []

    regs = pyg_data.get(2026,[])
    if regs:
        ing=sum(r['ing'] for r in regs); costo=sum(r['costo'] for r in regs)
        gasto=sum(r['gastos'] for r in regs); res=sum(r['res'] for r in regs)
        nom=sum(r['nom'] for r in regs); inst=sum(r['inst'] for r in regs)
        viaj=sum(r['viaj'] for r in regs); gest=sum(r['gest'] for r in regs)
        serv=sum(r['serv'] for r in regs); otros=sum(r['otros'] for r in regs)
        imp=res*0.3334 if res>0 else 0
        meses=sorted(set(r['month'] for r in regs))
        print(f"\nPyG 2026 (meses {meses}):")
        print(f"  Ingresos:  ${ing:>12,.2f}")
        print(f"  Costos:    ${costo:>12,.2f}")
        print(f"  Gastos:    ${gasto:>12,.2f}")
        print(f"    · Nómina:     ${nom:>10,.2f}  ({nom/gasto*100:.1f}%)")
        print(f"    · Servicios:  ${serv:>10,.2f}  ({serv/gasto*100:.1f}%)")
        print(f"    · Instalac.:  ${inst:>10,.2f}  ({inst/gasto*100:.1f}%)")
        print(f"    · Viajes:     ${viaj:>10,.2f}  ({viaj/gasto*100:.1f}%)")
        print(f"    · Gestión:    ${gest:>10,.2f}  ({gest/gasto*100:.1f}%)")
        print(f"    · Otros:      ${otros:>10,.2f}  ({otros/gasto*100:.1f}%)")
        print(f"  Resultado: ${res:>12,.2f}")
        print(f"  Neta:      ${res-imp:>12,.2f}")

        # ── Validaciones de calidad ──────────────────────────────────────
        # 1. Cierre aritmético: gastos = nom+inst+viaj+gest+serv+otros
        gastos_check = round(nom+inst+viaj+gest+serv+otros, 2)
        if abs(gastos_check - round(gasto, 2)) > 0.05:
            alertas.append(f"⚠️  GASTOS no cuadra: subgrupos={gastos_check:,.2f} vs total={gasto:,.2f} (diff={gastos_check-gasto:+,.2f})")

        # 2. Cierre aritmético: res = ing - costo - gastos
        res_check = round(ing - costo - gasto, 2)
        if abs(res_check - round(res, 2)) > 0.05:
            alertas.append(f"⚠️  RESULTADO no cuadra: ing-costo-gastos={res_check:,.2f} vs res={res:,.2f}")

        # 3. Alerta si 'otros' > 15% de gastos totales (indica cuentas sin clasificar)
        pct_otros = otros / gasto * 100 if gasto > 0 else 0
        if pct_otros > 15:
            alertas.append(
                f"⚠️  CLASIFICACIÓN: 'otros' = {pct_otros:.1f}% de gastos (${otros:,.2f}). "
                f"Revisar gasto_grupo() — probable cuentas sin keyword match."
            )

        # 4. Alerta si 'serv' < 10% de gastos (para empresa de comunicación es anómalo)
        pct_serv = serv / gasto * 100 if gasto > 0 else 0
        if pct_serv < 10:
            alertas.append(
                f"⚠️  CLASIFICACIÓN: 'serv' = {pct_serv:.1f}% de gastos (${serv:,.2f}). "
                f"Inusualmente bajo para empresa de servicios — revisar keywords."
            )

        # 5. Alerta si nom > 60% de gastos (puede indicar honorarios mal clasificados)
        pct_nom = nom / gasto * 100 if gasto > 0 else 0
        if pct_nom > 60:
            alertas.append(
                f"⚠️  NÓMINA: representa {pct_nom:.1f}% de gastos (${nom:,.2f}). "
                f"Verificar que honorarios sin dependencia no estén en nom."
            )

        # 6. Alerta si margen neto < -5% o > 30% (outliers que merecen revisión)
        margen = (res-imp)/ing*100 if ing > 0 else 0
        if margen < -5:
            alertas.append(f"⚠️  MARGEN NETO negativo: {margen:.1f}% — verificar datos.")
        elif margen > 30:
            alertas.append(f"⚠️  MARGEN NETO muy alto: {margen:.1f}% — verificar datos.")

    regs_f = fac_data.get(2026,[])
    if regs_f:
        total     = sum(r['valor'] for r in regs_f)
        reembolsos= sum(r['valor'] for r in regs_f if r.get('reembolso'))
        fac_pura  = total - reembolsos
        meses_f   = sorted(set(r['mes'] for r in regs_f))
        alertas_f = [r for r in regs_f if r.get('alerta')]
        print(f"\nFacturación 2026 (meses {meses_f}):")
        print(f"  Facturación:   ${fac_pura:>12,.2f}")
        print(f"  Reembolsos:    ${reembolsos:>12,.2f}")
        print(f"  Total:         ${total:>12,.2f}")
        if alertas_f:
            print(f"  ⚠️  {len(alertas_f)} registro(s) sin cliente identificado:")
            seen = set()
            for r in alertas_f:
                k = (r['cliente'], r['mes'])
                if k not in seen:
                    seen.add(k)
                    print(f"     {r['cliente']}  mes={r['mes']}  ${r['valor']:,.2f}")

        # Validación: facturación vs ingresos PyG no debería diferir más de 20%
        if regs and ing > 0:
            diff_pct = abs(total - ing) / ing * 100
            if diff_pct > 20:
                alertas.append(
                    f"⚠️  Facturación (${total:,.0f}) difiere {diff_pct:.1f}% de ingresos PyG (${ing:,.0f}). "
                    f"Verificar si hay diferencias de período o criterio."
                )

    if cxc_data:
        total=sum(r['valor'] for r in cxc_data)
        venc=sum(r['valor'] for r in cxc_data if r['estado']=='Vencida')
        print(f"\nCxC ({len(cxc_data)} registros):")
        print(f"  Total:     ${total:>12,.2f}")
        print(f"  Vencidas:  ${venc:>12,.2f}")
        pct_venc = venc/total*100 if total > 0 else 0
        if pct_venc > 50:
            alertas.append(f"⚠️  CxC: {pct_venc:.0f}% vencidas — revisar gestión de cobro.")

    if saldo_data:
        si_t=sum(r['si'] for r in saldo_data)
        sf_t=sum(r['sf'] for r in saldo_data)
        print(f"\nBancos:")
        print(f"  SI total:  ${si_t:>12,.2f}")
        print(f"  SF total:  ${sf_t:>12,.2f}")
        print(f"  Variación: ${sf_t-si_t:>12,.2f}")
        for r in saldo_data:
            print(f"    {r['banco']:<35} SI={r['si']:>10,.2f}  SF={r['sf']:>10,.2f}")

        # Validación: saldo negativo en cualquier cuenta bancaria
        for r in saldo_data:
            if r['sf'] < 0:
                alertas.append(f"⚠️  BANCOS: saldo negativo en '{r['banco']}': ${r['sf']:,.2f}")

    if mexico:
        total=sum(r['egr'] for r in mexico)
        print(f"\nFlujo México:")
        for r in mexico:
            print(f"  {MESES[r['mes']]}: ${r['egr']:,.2f}")
        print(f"  Total:     ${total:>12,.2f}")

    # ── Imprimir alertas ─────────────────────────────────────────────────
    print(f"\n{sep}")
    if alertas:
        print(f"⚠️  ALERTAS ({len(alertas)}):")
        for a in alertas:
            print(f"  {a}")
    else:
        print("✅ Sin alertas — todos los controles pasan.")
    print(sep)


# ══════════════════════════════════════════════════════════════════════════════
# 7. Publicar en GitHub
# ══════════════════════════════════════════════════════════════════════════════
def publicar_github():
    import subprocess
    print("\n🚀 Publicando en GitHub Pages...")

    fecha = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')

    msg = f'Actualizacion {fecha}'
    escenarios_path = BASE_DIR / 'escenarios.json'
    add_files = ['index.html']
    if escenarios_path.exists():
        add_files.append('escenarios.json')
    comandos = [
        ['git', 'add'] + add_files,
        ['git', 'commit', '-m', msg],
        ['git', 'push'],
    ]

    for cmd in comandos:
        resultado = subprocess.run(
            cmd, cwd=BASE_DIR,
            capture_output=True, text=True
        )
        salida = resultado.stdout + resultado.stderr
        if resultado.returncode != 0:
            if any(x in salida for x in ['nothing to commit', 'nothing added', 'no changes']):
                print("  ℹ️  Sin cambios nuevos que publicar")
                return
            print(f"  ❌ Error en '{' '.join(cmd)}':")
            print(f"     {salida.strip()}")
            return

    print("  ✅ Dashboard publicado en:")
    print("     https://kw-815.github.io/Dashboard-Financiero/\n")


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════
def main():
    if ARGS.debug_sp:
        debug_sharepoint()
        return

    modo = "LOCAL 📁" if ARGS.local else "SHAREPOINT ☁️"
    print(f"\n🔄 Dashboard Financiero Keyword — Actualizando... [{modo}]\n")

    if not INDEX_PATH.exists():
        print(f"❌ No se encontró index.html en {BASE_DIR}")
        return

    # ── Cargar archivos
    print(f"{'─'*40}")
    if ARGS.local:
        print("Cargando archivos locales...")
        contenidos = cargar_local()
    else:
        print("Conectando a SharePoint...")
        try:
            contenidos = cargar_sharepoint()
        except Exception as e:
            print(f"\n⚠️  Error SharePoint: {e}")
            print("\n  Intentando con archivos locales como fallback...")
            contenidos = cargar_local()

    if not contenidos:
        print("❌ No se encontraron archivos. Abortando.")
        return

    print(f"{'─'*40}\n")

    # ── Procesar
    print("1/5 Procesando PyG...")
    pyg_data = procesar_pyg(contenidos)

    print("\n2/5 Procesando Facturación...")
    fac_data = procesar_facturacion(contenidos)

    print("\n3/5 Procesando CxC...")
    cxc_data = procesar_cxc(contenidos)

    print("\n4/5 Procesando Bancos...")
    saldo_data, monthly, mexico = procesar_bancos(contenidos)

    print("\n5/5 Actualizando index.html...")
    actualizar_html(pyg_data, fac_data, cxc_data, saldo_data, monthly, mexico)

    imprimir_resumen(pyg_data, fac_data, cxc_data, saldo_data, monthly, mexico)

    # ── Publicar en GitHub automáticamente
    if not ARGS.local:
        publicar_github()

    print("\n✅ Listo. Abre index.html en tu navegador para verificar.\n")


if __name__ == '__main__':
    main()
